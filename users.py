import openpyxl


# open file with all employees
while True:
    filename = input("Name of the file with employees data (with .xlsx): ")
    try:
        wb = openpyxl.load_workbook(filename, data_only=True)
        break
    except:
        print("no such file in directory, please check the name")
ws = wb.active                                              # current worksheet


empl_duplicates = []
employees = dict()
for i in range(1, ws.max_row+1):                            # iterate through the current ws
    surname = ws.cell(row=i, column=1).value.strip()            
    empl_data = []                                          # create empty list for employee's data
    if surname in employees:                                # if the surname already exists in employees dict
        for j in range(1, 9):                                   # create a list of this one employee's data
            empl_data.append(ws.cell(row=i, column=j).value)
        empl_duplicates.append(empl_data)                       # and append to list empl_duplicates
    else:                                                   # else
        for j in range(2, 9):                                   # read employee's data from column 2 to 8 incl.
            empl_data.append(ws.cell(row=i, column=j).value)    # and append them to employee's list
        employees[surname] = empl_data                          # add employee's data do employess dist


while True:
    filename = input("Name of the file with user's licences (with .xlsx): ")
    try:
        wb_user = openpyxl.load_workbook(filename, data_only=True)
        break
    except:
        print("no such file in directory, please check the name")
ws_user = wb_user.active
ws_error = wb_user.create_sheet("error")  
errors = set()                                                  # store employees names which were not found 

# create header
column = 6                                                      # starting position
for name in ("company name", "profit center", "BU"):
    ws_user.cell(row=1, column=column).value = name
    column += 1

for i in range(2, ws_user.max_row+1):                           # iterate through column D (surname)
    actual_cell = ws_user.cell(row=i, column=4).value
    try:
        if employees[actual_cell]:                              # if you find the employee's last name in employees dictionary
            if employees[actual_cell][0].strip() == ws_user.cell(row=i, column=5).value:    # check if the first name is the same
                counter = 6                                     # add from column 6 info about the employee
                for j in (2, 5, 6):                             # index of empl_data stored in employees dict (company, profit center, BU)
                    ws_user.cell(row=i, column=counter).value = employees[actual_cell][j]
                    counter += 1
            else:                                               # if the first name is different
                for empl in empl_duplicates:                    # iterate through list of dupliates employees
                    if empl[0] == actual_cell and empl[1].strip() == ws_user.cell(row=i, column=5).value: # check the first and last name
                        counter = 6                             # add from column 6 info about the employee
                        for j in (3, 6, 7):                     # index of employee's data stored in a list of empl_duplicates list
                            ws_user.cell(row=i, column=counter).value = empl[j]
                            counter += 1
    except:                                                     # if employee is not in dictionary, copy the name to sheet error
        errors.add((actual_cell, ws_user.cell(row=i, column=5).value))
for error in errors:
    ws_error.append(error)                                               

for i in range(2, ws_user.max_row+1):                           # iterate through column F
    actual_cell = ws_user.cell(row=i, column=6).value
    if not actual_cell:                                         # if the cell is empty
        try:
            if employees[ws_user.cell(row=i, column=4).value]:  # try to find the surname
                counter = 6
                for j in (2, 5, 6):                             # index of empl_data stored in employees dict
                    ws_user.cell(row=i, column=counter).value = employees[ws_user.cell(row=i, column=4).value][j]
                    counter += 1
        except:
            pass
            

for i in range(2, ws_user.max_row+1):                           # iterate throug column F
    actual_cell = ws_user.cell(row=i, column=6).value
    if actual_cell:                                             # ifthe cell is not empty, then convert the sting into list (delimeter is " - "),
        ws_user.cell(row=i, column=6).value = actual_cell.split(" - ")[1]   # choose the 2nd element from the list (organization name)
        

wb_user.save("users_result.xlsx")

print("*************************************************************************************")
print("File users_result.xlsx has been created. You can check the blank cells in column F")
print("If you wish, you can add the dimension manually or leave it blank.")
print("Or copy the names from sheet error to file with all employees, add dimensions and rerun python users.py")
print("To calculate costs per cost centrum run python3.6 user_cost.py")
print("*************************************************************************************")

"""
co shodná příjmení? - pokud nekontroluji jméno, budou přiřazené špatné dimenze
budu-li zároveň kontrolovat i křestní jméno, tak to bývá občas vyplněné špatně
a nedotáhnou se tak žádné dimmenze
"""